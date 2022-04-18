# -*- encoding: utf-8 -*-
from flask import Flask, request
from flask_restful import Resource, Api
import requests
import json
import os
import numpy as np
import pickle
from SPARQLWrapper import SPARQLWrapper, JSON
import pandas as pd
from hazm import *
from openpyxl import load_workbook
from io import BytesIO
import urllib

app = Flask(__name__)
api = Api(app)

def callApi(url, data, tokenKey):
    headers = {
        'Content-Type': "application/json",
        'Authorization': "Bearer " + tokenKey,
        'Cache-Control': "no-cache"
    }
    response = requests.request("POST", url, data=data.encode("utf-8"), headers=headers)
#     print(response)
    return response.text

def mergeTwoListWithSomeEqualElements(list1, list2):
    mergedList = []
    for i in range(0, len(list1)):
        if(list1[i] in list2):
            mergedList.append(list1[i])
    
    if(len(mergedList) > 0):
        for i in range(0, len(list1)):
            if(list1[i] not in mergedList):
                mergedList.append(list1[i])
        for i in range(0, len(list2)):
            if(list2[i] not in mergedList):
                mergedList.append(list2[i])
    return mergedList

def getStemmOfArray(StemmerUrl, tokenKey, synonyms):
    payload = ""
    for i in range(0, len(synonyms)):
        payload += synonyms[i]
        if(i<len(synonyms)-1):
            payload += " ، "

    payload = u'"' + payload + '"'
#     print("payload = ", payload)
    stemmerResult = callApi(StemmerUrl, payload, tokenKey)
#     print(stemmerResult)
    resultArray = stemmerResult.split('، ')
    return resultArray

def removeUnnecessaryCharacters(sentence):
    unnecessaryCharacters = ['»', '«', '\"', '\"', ')', '(', '+', '-', '،', '.', '؟', '!', '؛', '...', '…', '[', ']', '/', '=', ':', ',', 'ٔ', '  ']
    for i in range (0, len(unnecessaryCharacters)):
        sentence = sentence.replace(unnecessaryCharacters[i], '')
        
    sentence = ' '.join(sentence.split())

    return sentence

def isRelationAlreadyExist(validRelations, newValidRelation):
    for i in range(0, len(validRelations)):
        if(newValidRelation['head'] == validRelations[i]['head'] and 
        newValidRelation['tail'] == validRelations[i]['tail'] and
        newValidRelation['relation'] == validRelations[i]['relation']):
            return True
    return False

class CheckSentenceRelations(Resource):

    @staticmethod
    def get():
        sentence = request.args.get('sentence', '')
        textMiningIrApiKey = request.args.get('textMiningIrApiKey', '')

        headers = {
            'Content-Type': "application/json",
            'Cache-Control': "no-cache"
        }

        ##################### Get Token by Api Key ##########################
        baseTextMiningUrl = "http://api.text-mining.ir/api/"
        url = baseTextMiningUrl + "Token/GetToken"
        querystring = {"apikey":textMiningIrApiKey}
        response = requests.request("GET", url, params=querystring)
        responseData = json.loads(response.text)
        try:
            tokenKey = responseData['token']
        except Exception as e:
            return "Token Key is not valid or expired!"


        StemmerUrl =  baseTextMiningUrl + "Stemmer/LemmatizeText2Text"
        farsBaseEntityLinkingURL = "http://farsbase.net:8099/proxy/raw/rest/v1/raw/FKGfy"
        sparqlEndpoint = "http://farsbase.net:8890/sparql"
        sparql = SPARQLWrapper(sparqlEndpoint)

        sentence = removeUnnecessaryCharacters(sentence)
        # print("sentence = ", sentence)
        requestSentence = {"text": sentence}
        headers = {'content-type': 'application/json'}
        response = requests.post(farsBaseEntityLinkingURL, data=json.dumps(requestSentence), headers=headers)
        d = json.loads(response.text)
        
        sentenceWords = []
        sentenceIris = []
        for i in range(0, len(d[0])):
            iris = []
            try:
        #         print(d[0][i])
                if('iri' in d[0][i]['resource']):
        #             if(len(d[0][i]['resource']['classes']) > 1 or ((len(d[0][i]['resource']['classes']) == 1) and ('http://fkg.iust.ac.ir/ontology/Thing' not in d[0][i]['resource']['classes'] or len(d[0][i]['ambiguities']) > 0 )
                    if(len(d[0][i]['resource']['classes']) > 1 or ((len(d[0][i]['resource']['classes']) == 1) and ('http://fkg.iust.ac.ir/ontology/Thing' not in d[0][i]['resource']['classes']))):
        #                 print(d[0][i])
                        iris.append(d[0][i]['resource']['iri'])
                        for k in range (0, len(d[0][i]['ambiguities'])):
                            iris.append(d[0][i]['ambiguities'][k]['iri'])
                        sentenceWords.append(d[0][i]['word'])
                        sentenceIris.append(iris)
            except Exception as e:
                pass
        # print("sentenceIris = ", sentenceIris)        
        newSentenceWords = []
        newSentenceIris = []
        for i in range (0, len(sentenceWords)):
            if(len(newSentenceIris) > 0):
                newIris = mergeTwoListWithSomeEqualElements(newSentenceIris[len(newSentenceIris)-1], sentenceIris[i])
                if(len(newIris) > 0):
                    newSentenceIris[len(newSentenceIris)-1] = newIris
                    newSentenceWords[len(newSentenceWords)-1] = newSentenceWords[len(newSentenceWords)-1] + " " + sentenceWords[i]
                else:
                    newSentenceIris.append(sentenceIris[i])
                    newSentenceWords.append(sentenceWords[i])
            else:
                newSentenceIris.append(sentenceIris[i])
                newSentenceWords.append(sentenceWords[i])

        # print("newSentenceIris = ", newSentenceIris)    
        allFoundedRelationsInSentence = []
        for i  in range (0, len(newSentenceIris)):
            for j in range (0, len(newSentenceIris[i])):
                for k  in range (0, len(newSentenceIris)):
                    if(k != i):
                        for l  in range (0, len(newSentenceIris[k])):
                            stringQuery = """select ?p WHERE {
                            <""" + newSentenceIris[i][j] + """> ?p <""" + newSentenceIris[k][l] + """>
                            }"""
                            # print("stringQuery = ", stringQuery)
                            sparql.setQuery(stringQuery)
                            sparql.setReturnFormat(JSON)
                            results = sparql.query().convert()
                            # print("query Result = ", results)
                            if(len(results['results']['bindings']) > 0):
                                for m in range (0, len(results['results']['bindings'])):
                                    foundedRelation = {}
                                    foundedRelation['sentence'] = sentence

                                    head = {}
                                    head['word'] = newSentenceWords[i]
                                    head['id'] = newSentenceIris[i][j]
                                    foundedRelation['head'] = head

                                    tail = {}
                                    tail['word'] = newSentenceWords[k]
                                    tail['id'] = newSentenceIris[k][l]
                                    foundedRelation['tail'] = tail
                                    
                                    foundedRelation['relation'] = results['results']['bindings'][m]['p']['value']
                                    
                                    if (not (foundedRelation in allFoundedRelationsInSentence)):
                                        allFoundedRelationsInSentence.append(foundedRelation)
                                    # print(foundedRelation)
                                    
                            stringQuery = """SELECT ?p WHERE {
                                <""" + newSentenceIris[i][j] + """> <http://fkg.iust.ac.ir/ontology/relatedPredicates> ?o.
                                ?o ?p <""" + newSentenceIris[k][l] + """>
                                }"""

                            sparql.setQuery(stringQuery)
                            sparql.setReturnFormat(JSON)
                            results = sparql.query().convert()
                            for m in range (0, len(results['results']['bindings'])):
                                    foundedRelation = {}
                                    foundedRelation['sentence'] = sentence

                                    head = {}
                                    head['word'] = newSentenceWords[i]
                                    head['id'] = newSentenceIris[i][j]
                                    foundedRelation['head'] = head

                                    tail = {}
                                    tail['word'] = newSentenceWords[k]
                                    tail['id'] = newSentenceIris[k][l]
                                    foundedRelation['tail'] = tail
                                    
                                    foundedRelation['relation'] = results['results']['bindings'][m]['p']['value']
                                    
                                    if (not (foundedRelation in allFoundedRelationsInSentence)):
                                        allFoundedRelationsInSentence.append(foundedRelation)
                                    # print(foundedRelation)



        # file = urllib.request.urlopen('Properties_refined_two_side_entity_filterd.xlsx').read()
        # dfs = load_workbook(filename = 'Properties_refined_two_side_entity_filterd.xlsx')

        # dfs = pd.read_excel("Properties_refined_two_side_entity_filterd.xlsx", sheet_name="Sheet1", engine='openpyxl')
        # relations = []
        # for i in range (0, len(dfs)):
            # relations.append(dfs['URL'][i])

        dfs = load_workbook(filename = 'Properties_refined_two_side_entity_filterd.xlsx')

        # print(dfs['Sheet1'].cell(row=4, column=2).value)

        relations = []
        for i in range (2, 214):
            relations.append(dfs['Sheet1'].cell(row=i, column=1).value)

        synonymsArray = []
        for i in range (2, 214):
            ss = ""
            if (not(dfs['Sheet1'].cell(row=i, column=5).value == None)):
                ss += dfs['Sheet1'].cell(row=i, column=5).value
            if (not(dfs['Sheet1'].cell(row=i, column=6).value == None)):
                ss += " , " + dfs['Sheet1'].cell(row=i, column=6).value
            synonymsArray.append(ss)

        data = allFoundedRelationsInSentence
        # print("allFoundedRelationsInSentence = ", allFoundedRelationsInSentence)
        normalizer = Normalizer()
        notInExecelCount = 0
        validRelations = []
        for i in range (0, len(data)):
            try:
                sentence = data[i]['sentence']
                firstEntity = data[i]['head']['id']
                firstWord = data[i]['head']['word']
                secondWord = data[i]['tail']['word']
                secondEntity = data[i]['tail']['id']
                relation = data[i]['relation']
                if (relation in relations):
                    sentence = removeUnnecessaryCharacters(sentence)
                    if (firstWord == secondWord):
                        break
                    if (firstEntity == secondEntity):
                        break

                    if (not (synonymsArray[relations.index(relation)] == None)):
                        synonyms = synonymsArray[relations.index(relation)].replace(u'\xa0', u' ').replace(u' , ', u',').replace(u', ', u',').replace(u' ,', u',').replace(u'\u200c', u' ').split(',')

                        splitedSentence = sentence.split(' ')
                        for synonymIndex in range (0, len(synonyms)):
                            for i in range (0, len(splitedSentence)):
                                if(not(normalizer.normalize(synonyms[synonymIndex]) == '') and normalizer.normalize(synonyms[synonymIndex]) == splitedSentence[i]):
                                    newValidRelation = {}
                                    newValidRelation['sentence'] = sentence
                                    firstEn = {}
                                    firstEn['word'] = firstWord
                                    firstEn['start'] = sentence.find(firstWord)
                                    firstEn['length'] = len(firstWord)
                                    firstEn['id'] = firstEntity
                                    secondEn = {}
                                    secondEn['word'] = secondWord
                                    secondEn['start'] = sentence.find(secondWord)
                                    secondEn['length'] = len(secondWord)
                                    secondEn['id'] = secondEntity
                                    newValidRelation['head'] = firstEn
                                    newValidRelation['tail'] = secondEn
                                    newValidRelation['relation'] = relation
                                    newValidRelation['why'] = "No root : " + normalizer.normalize(synonyms[synonymIndex]) + " == " + splitedSentence[i]
                                    if(not isRelationAlreadyExist(validRelations, newValidRelation)):
                                        validRelations.append(newValidRelation)
                                    break

                        stemmerResultOnSynonyms = getStemmOfArray(StemmerUrl, tokenKey, synonyms)
                        for stemmerResultIndex in range (0, len(stemmerResultOnSynonyms)):
                            for i in range (0, len(splitedSentence)):
                                if(not(normalizer.normalize(stemmerResultOnSynonyms[stemmerResultIndex]) == '') and normalizer.normalize(stemmerResultOnSynonyms[stemmerResultIndex]) == splitedSentence[i]):
                                    newValidRelation = {}
                                    newValidRelation['sentence'] = sentence
                                    firstEn = {}
                                    firstEn['word'] = firstWord
                                    firstEn['start'] = sentence.find(firstWord)
                                    firstEn['length'] = len(firstWord)
                                    firstEn['id'] = firstEntity
                                    secondEn = {}
                                    secondEn['word'] = secondWord
                                    secondEn['start'] = sentence.find(secondWord)
                                    secondEn['length'] = len(secondWord)
                                    secondEn['id'] = secondEntity
                                    newValidRelation['head'] = firstEn
                                    newValidRelation['tail'] = secondEn
                                    newValidRelation['relation'] = relation
                                    newValidRelation['why'] = "Root of Synonym: " + normalizer.normalize(stemmerResultOnSynonyms[stemmerResultIndex]) + " == " + splitedSentence[i]
                                    if(not isRelationAlreadyExist(validRelations, newValidRelation)):
                                        validRelations.append(newValidRelation)
                                    break

                        payload = u'"' + sentence + '"'
                        sentenceStemmerResult = callApi(StemmerUrl, payload, tokenKey)
                        splitedSentenceStemmerResult = sentenceStemmerResult.split(' ')
                        for stemmerResultIndex in range (0, len(stemmerResultOnSynonyms)):
                            for i in range (0, len(splitedSentenceStemmerResult)):
                                if(not(normalizer.normalize(stemmerResultOnSynonyms[stemmerResultIndex]) == '') and normalizer.normalize(stemmerResultOnSynonyms[stemmerResultIndex]) == splitedSentenceStemmerResult[i]):
                                    newValidRelation = {}
                                    newValidRelation['sentence'] = sentence
                                    firstEn = {}
                                    firstEn['word'] = firstWord
                                    firstEn['start'] = sentence.find(firstWord)
                                    firstEn['length'] = len(firstWord)
                                    firstEn['id'] = firstEntity
                                    secondEn = {}
                                    secondEn['word'] = secondWord
                                    secondEn['start'] = sentence.find(secondWord)
                                    secondEn['length'] = len(secondWord)
                                    secondEn['id'] = secondEntity
                                    newValidRelation['head'] = firstEn
                                    newValidRelation['tail'] = secondEn
                                    newValidRelation['relation'] = relation
                                    newValidRelation['why'] = "Root of both: " + normalizer.normalize(stemmerResultOnSynonyms[stemmerResultIndex]) + " == " + splitedSentenceStemmerResult[i]
                                    if(not isRelationAlreadyExist(validRelations, newValidRelation)):
                                        validRelations.append(newValidRelation)
                                    break
                else:
                    notInExecelCount = notInExecelCount + 1
            except Exception as e:
                pass
                
        # print(validRelations)

        return validRelations

api.add_resource(CheckSentenceRelations, '/')

if __name__ == '__main__':
    app.run(host='127.0.0.1', port=9030)
